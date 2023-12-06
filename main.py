import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Alignment

def cadastrar_funcionario():
    nome = entry_nome.get()
    cargo = entry_cargo.get()

    if nome and cargo:
        adicionar_funcionario(nome, cargo)
        messagebox.showinfo("Sucesso", "Funcionário cadastrado e exportado para o Excel.")
    else:
        messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")

def adicionar_funcionario(nome, cargo):
    planilha = 'relatorio_funcionarios.xlsx'

    try:
        workbook = openpyxl.load_workbook(planilha)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Cargo'])
        header_row = sheet[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    sheet = workbook.active

    # Encontrar a próxima linha vazia
    proxima_linha = sheet.max_row + 1

    sheet.cell(row=proxima_linha, column=1, value=nome)
    sheet.cell(row=proxima_linha, column=2, value=cargo)

    workbook.save(planilha)

def limpar_dados():
    planilha = 'relatorio_funcionarios.xlsx'
    try:
        workbook = openpyxl.load_workbook(planilha)
        sheet = workbook.active

        # Limpa todas as linhas exceto a primeira (cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        workbook.save(planilha)
        messagebox.showinfo("Sucesso", "Dados limpos na planilha.")
    except FileNotFoundError:
        messagebox.showwarning("Aviso", "A planilha não foi encontrada.")

# Criar a janela principal
janela = tk.Tk()
janela.title("Cadastro de Funcionários")

# Obter dimensões da tela
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()

# Ajustar a largura e altura dos widgets conforme necessário
largura_widgets = int(largura_tela / 3)
altura_widgets = int(altura_tela / 5)

# Criar os widgets
label_nome = tk.Label(janela, text="Nome:")
entry_nome = tk.Entry(janela, width=int(largura_widgets / 20))

label_cargo = tk.Label(janela, text="Cargo:")
entry_cargo = tk.Entry(janela, width=int(largura_widgets / 20))

botao_cadastrar = tk.Button(janela, text="Cadastrar", command=cadastrar_funcionario, width=int(largura_widgets / 20))
botao_limpar = tk.Button(janela, text="Limpar Dados", command=limpar_dados, width=int(largura_widgets / 20))

# Posicionar os widgets na janela
label_nome.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
entry_nome.grid(row=0, column=1, padx=10, pady=5)

label_cargo.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_cargo.grid(row=1, column=1, padx=10, pady=5)

botao_cadastrar.grid(row=2, column=0, columnspan=2, pady=5)
botao_limpar.grid(row=3, column=0, columnspan=2, pady=5)

# Definir tamanho da janela
janela.geometry(f"{largura_widgets}x{altura_widgets}+{(largura_tela - largura_widgets)//2}+{(altura_tela - altura_widgets)//2}")

# Iniciar o loop principal
janela.mainloop()